"""Excel 解析器模块"""

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from models.enums import Channel, TaskStatus
from models.task import Task
from models.content import Content
from services.config_manager import get_config_manager
from .path_mapper import PathMapper, get_path_mapper

logger = logging.getLogger(__name__)

# 列名映射（支持多种命名方式）
COLUMN_MAPPINGS = {
    "content_code": ["文案编号", "产品编号", "产品编码", "内容编码", "编码", "code", "content_code", "产品代码"],
    "channel": ["发布位置", "发布渠道", "渠道", "channel", "类型"],
    "scheduled_time": ["排期时间", "发布时间", "时间", "scheduled_time", "schedule", "发送时间"],
    "product_name": ["产品名称", "产品", "名称", "product_name", "product"],
    "text": ["文案内容", "文案", "内容", "text", "content", "正文"],
    "image_paths": ["图片路径", "图片", "images", "image_paths", "图片地址", "文件夹路径"],
    "group_name": ["群名", "群名称", "群组", "group", "group_name"],
    "image_count": ["图片数", "图片数量", "image_count"],
    "category": ["分类", "类别", "category"],
    "product_link": ["产品链接", "链接", "product_link", "link"],
}

# 渠道值映射
# 注意：旧的 "群发"/"group" 默认归类为代理群
CHANNEL_VALUE_MAPPINGS = {
    "moment": ["moment", "朋友圈", "pyq", "moments"],
    "agent_group": ["agent_group", "代理群", "代理", "群发", "group", "qf", "群", "groups"],
    "customer_group": ["customer_group", "客户群", "客户"],
}

# 必需列 (scheduled_time 改为可选，可以后续在GUI中设置)
REQUIRED_COLUMNS = ["content_code", "channel"]


@dataclass
class ParseError:
    """解析错误"""
    row: int
    column: str
    value: Any
    message: str


@dataclass
class ParseWarning:
    """解析警告"""
    row: int
    column: str
    message: str


@dataclass
class ParseResult:
    """解析结果"""
    success: bool
    tasks: List[Task] = field(default_factory=list)
    contents: Dict[str, Content] = field(default_factory=dict)  # content_code -> Content
    errors: List[ParseError] = field(default_factory=list)
    warnings: List[ParseWarning] = field(default_factory=list)
    total_rows: int = 0
    valid_rows: int = 0

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0

    @property
    def has_warnings(self) -> bool:
        return len(self.warnings) > 0


class ExcelParser:
    """Excel 解析器"""

    def __init__(
        self,
        path_mapper: Optional[PathMapper] = None,
        strict_mode: bool = False,
    ):
        """
        初始化解析器

        Args:
            path_mapper: 路径映射器
            strict_mode: 严格模式（任何错误都中止解析）
        """
        if openpyxl is None:
            raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

        self.path_mapper = path_mapper or get_path_mapper()
        self.strict_mode = strict_mode

    def parse(self, file_path: str, sheet_name: Optional[str] = None) -> ParseResult:
        """
        解析 Excel 文件

        Args:
            file_path: Excel 文件路径
            sheet_name: 工作表名称（默认为第一个）

        Returns:
            ParseResult
        """
        result = ParseResult(success=True)
        path = Path(file_path)

        # 检查文件存在
        if not path.exists():
            result.success = False
            result.errors.append(ParseError(
                row=0, column="", value=file_path,
                message=f"文件不存在: {file_path}"
            ))
            return result

        # 检查文件扩展名
        if path.suffix.lower() not in (".xlsx", ".xls"):
            result.success = False
            result.errors.append(ParseError(
                row=0, column="", value=file_path,
                message=f"不支持的文件格式: {path.suffix}"
            ))
            return result

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # 选择工作表
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    result.success = False
                    result.errors.append(ParseError(
                        row=0, column="", value=sheet_name,
                        message=f"工作表不存在: {sheet_name}"
                    ))
                    return result
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # 解析表头
            column_map = self._parse_header(ws)

            # 检查必需列
            missing_columns = []
            for col in REQUIRED_COLUMNS:
                if col not in column_map:
                    missing_columns.append(col)

            if missing_columns:
                result.success = False
                result.errors.append(ParseError(
                    row=1, column="", value="",
                    message=f"缺少必需列: {', '.join(missing_columns)}"
                ))
                return result

            # 解析数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                result.total_rows += 1

                # 跳过空行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # 解析行数据
                row_result = self._parse_row(row, row_idx, column_map)

                if row_result["errors"]:
                    result.errors.extend(row_result["errors"])
                    if self.strict_mode:
                        result.success = False
                        return result
                else:
                    task = row_result["task"]
                    content = row_result["content"]

                    if task:
                        task.source_folder = str(path.parent)
                        result.tasks.append(task)
                        result.valid_rows += 1

                    if content and content.content_code:
                        # 合并同一 content_code 的内容
                        if content.content_code in result.contents:
                            existing = result.contents[content.content_code]
                            # 合并图片
                            for img in content.image_paths:
                                if img not in existing.image_paths:
                                    existing.image_paths.append(img)
                            # 如果新内容有文案且旧内容没有，则更新
                            if content.text and not existing.text:
                                existing.text = content.text
                        else:
                            result.contents[content.content_code] = content

                if row_result["warnings"]:
                    result.warnings.extend(row_result["warnings"])

            wb.close()

        except Exception as e:
            result.success = False
            result.errors.append(ParseError(
                row=0, column="", value="",
                message=f"解析 Excel 失败: {e}"
            ))
            logger.exception(f"解析 Excel 失败: {file_path}")

        return result

    def _parse_header(self, ws) -> Dict[str, int]:
        """
        解析表头，返回列名到列索引的映射

        Returns:
            {标准列名: 列索引}
        """
        column_map = {}
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue

            cell_str = str(cell_value).strip().lower()

            # 查找匹配的标准列名
            for std_name, aliases in COLUMN_MAPPINGS.items():
                if any(alias.lower() == cell_str for alias in aliases):
                    column_map[std_name] = col_idx
                    break

        return column_map

    def _parse_row(
        self,
        row: tuple,
        row_idx: int,
        column_map: Dict[str, int],
    ) -> Dict[str, Any]:
        """解析单行数据"""
        errors: List[ParseError] = []
        warnings: List[ParseWarning] = []

        def get_cell(col_name: str) -> Any:
            if col_name in column_map:
                idx = column_map[col_name]
                if idx < len(row):
                    return row[idx]
            return None

        # 获取单元格值
        content_code = get_cell("content_code")
        channel_raw = get_cell("channel")
        scheduled_time_raw = get_cell("scheduled_time")
        product_name = get_cell("product_name")
        category = get_cell("category")
        text = get_cell("text")
        image_paths_raw = get_cell("image_paths")
        group_name = get_cell("group_name")
        product_link = get_cell("product_link")

        # 验证必需字段
        if not content_code or str(content_code).strip() == "":
            errors.append(ParseError(
                row=row_idx, column="content_code", value=content_code,
                message="产品编码不能为空"
            ))

        # 解析渠道（使用包含匹配，支持自定义渠道）
        channel = self._parse_channel(channel_raw)
        if channel is None:
            # 渠道匹配失败，跳过该行（返回空结果，不报错）
            warnings.append(ParseWarning(
                row=row_idx, column="channel",
                message=f"渠道「{channel_raw}」未匹配到任何渠道，已跳过该行"
            ))
            return {"task": None, "content": None, "errors": [], "warnings": warnings}

        # 解析时间（可选字段）
        scheduled_time = None
        if scheduled_time_raw:
            scheduled_time = self._parse_datetime(scheduled_time_raw)
            if scheduled_time is None:
                warnings.append(ParseWarning(
                    row=row_idx, column="scheduled_time",
                    message=f"无效的时间格式: {scheduled_time_raw}，任务将设为待排期状态"
                ))

        # 解析图片路径
        image_paths = self._parse_image_paths(image_paths_raw)

        # 验证图片路径
        invalid_paths = []
        for img_path in image_paths:
            is_valid, error = self.path_mapper.validate_path(img_path)
            if not is_valid:
                invalid_paths.append(img_path)
                warnings.append(ParseWarning(
                    row=row_idx, column="image_paths",
                    message=f"图片路径无效: {img_path}"
                ))

        # 过滤掉无效路径
        valid_image_paths = [p for p in image_paths if p not in invalid_paths]

        # 群发渠道建议填写群名（自定义渠道也是群发渠道）
        is_group = Channel.is_group_channel(channel) if isinstance(channel, Channel) else Channel.is_custom_channel(channel)
        if is_group and not group_name:
            warnings.append(ParseWarning(
                row=row_idx, column="group_name",
                message="群发渠道建议填写群名"
            ))

        # 如果有错误，不创建对象
        if errors:
            return {"task": None, "content": None, "errors": errors, "warnings": warnings}

        # 创建 Task
        task = Task(
            content_code=str(content_code).strip(),
            product_name=str(product_name).strip() if product_name else "",
            category=str(category).strip() if category else "",
            product_link=str(product_link).strip() if product_link else "",
            text=str(text).strip() if text else "",
            channel=channel,
            group_name=str(group_name).strip() if group_name else None,
            status=TaskStatus.pending,
            scheduled_time=scheduled_time,
        )

        # 创建 Content
        content = Content(
            content_code=str(content_code).strip(),
            text=str(text).strip() if text else "",
            image_paths=valid_image_paths,
            channel=channel,
            product_link=str(product_link).strip() if product_link else "",
            product_name=str(product_name).strip() if product_name else "",
            category=str(category).strip() if category else "",
        )

        return {
            "task": task,
            "content": content,
            "errors": errors,
            "warnings": warnings,
        }

    def _parse_channel(self, value: Any):
        """
        解析渠道值（支持内置渠道和自定义渠道，使用包含匹配）

        匹配顺序：
        1. 完全匹配（优先级最高）
        2. 包含匹配（按匹配长度排序，更精确的优先）

        Returns:
            Channel枚举 或 自定义渠道ID字符串 或 None
        """
        if value is None:
            return None

        value_str = str(value).strip()
        value_lower = value_str.lower()

        # 收集所有候选匹配: (渠道, 匹配的别名/名称, 是否完全匹配)
        candidates = []

        # 1. 检查内置渠道
        for channel, aliases in CHANNEL_VALUE_MAPPINGS.items():
            for alias in aliases:
                alias_lower = alias.lower()
                # 完全匹配
                if value_lower == alias_lower:
                    return Channel(channel)  # 完全匹配直接返回
                # 包含匹配：value包含alias 或 alias包含value
                if alias_lower in value_lower or value_lower in alias_lower:
                    candidates.append((Channel(channel), alias, len(alias)))

        # 2. 检查自定义渠道
        config = get_config_manager()
        custom_channels = config.get_custom_channels() or {}

        for channel_id, channel_config in custom_channels.items():
            channel_name = channel_config.get("name", "")
            if not channel_name:
                continue
            channel_name_lower = channel_name.lower()
            # 完全匹配
            if value_lower == channel_name_lower:
                return channel_id  # 完全匹配直接返回
            # 包含匹配
            if channel_name_lower in value_lower or value_lower in channel_name_lower:
                candidates.append((channel_id, channel_name, len(channel_name)))

        # 3. 如果有候选，选择匹配名称最长的（更精确）
        if candidates:
            # 按匹配长度降序排序
            candidates.sort(key=lambda x: x[2], reverse=True)
            return candidates[0][0]

        return None

    def _parse_datetime(self, value: Any) -> Optional[datetime]:
        """解析日期时间"""
        if value is None:
            return None

        # 如果已经是 datetime 对象
        if isinstance(value, datetime):
            return value

        value_str = str(value).strip()

        # 尝试多种日期格式
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y.%m.%d %H:%M:%S",
            "%Y.%m.%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value_str, fmt)
            except ValueError:
                continue

        return None

    def _parse_image_paths(self, value: Any) -> List[str]:
        """解析图片路径（支持文件夹路径，自动扫描图片）"""
        if value is None:
            return []

        value_str = str(value).strip()
        if not value_str:
            return []

        # 支持的图片格式
        image_extensions = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}

        # 使用路径映射器分割路径
        paths = self.path_mapper.split_paths(value_str)

        result = []
        for p in paths:
            if not p:
                continue

            normalized = self.path_mapper.normalize_path(p)
            path_obj = Path(normalized)

            # 如果是文件夹，扫描其中的图片
            if path_obj.is_dir():
                try:
                    for img_file in sorted(path_obj.iterdir()):
                        if img_file.is_file() and img_file.suffix.lower() in image_extensions:
                            result.append(str(img_file))
                except Exception as e:
                    logger.warning(f"扫描图片文件夹失败: {normalized}, 错误: {e}")
            # 如果是文件，直接添加
            elif path_obj.is_file():
                result.append(normalized)
            # 如果路径不存在但看起来像文件，也添加（后续验证会处理）
            elif path_obj.suffix.lower() in image_extensions:
                result.append(normalized)

        return result

    def validate_file(self, file_path: str) -> Tuple[bool, List[str]]:
        """
        快速验证 Excel 文件格式

        Returns:
            (是否有效, 错误列表)
        """
        errors = []
        path = Path(file_path)

        if not path.exists():
            return False, [f"文件不存在: {file_path}"]

        if path.suffix.lower() not in (".xlsx", ".xls"):
            return False, [f"不支持的文件格式: {path.suffix}"]

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # 检查表头
            column_map = self._parse_header(ws)

            for col in REQUIRED_COLUMNS:
                if col not in column_map:
                    errors.append(f"缺少必需列: {col}")

            wb.close()

        except Exception as e:
            return False, [f"无法读取文件: {e}"]

        return len(errors) == 0, errors


def parse_excel(file_path: str, sheet_name: Optional[str] = None) -> ParseResult:
    """便捷函数：解析 Excel 文件"""
    parser = ExcelParser()
    return parser.parse(file_path, sheet_name)


def validate_excel(file_path: str) -> Tuple[bool, List[str]]:
    """便捷函数：验证 Excel 文件"""
    parser = ExcelParser()
    return parser.validate_file(file_path)


def parse_folder(folder_path: str) -> ParseResult:
    """
    从文件夹解析任务和图片

    流程:
    1. 查找文件名包含"汇总"的 Excel 文件
    2. 解析 Excel 获取任务列表和图片数
    3. 根据 image_count 精确匹配图片文件

    图片文件名格式: 文案编号 (1).jpg, 文案编号 (2).jpg, ...

    Args:
        folder_path: 素材文件夹路径

    Returns:
        ParseResult: 解析结果
    """
    folder = Path(folder_path)

    if not folder.exists() or not folder.is_dir():
        return ParseResult(
            success=False,
            errors=[ParseError(row=0, column="", message=f"文件夹不存在: {folder_path}")]
        )

    # 1. 查找包含"汇总"的 Excel 文件
    excel_file = _find_excel_file(folder)
    if not excel_file:
        return ParseResult(
            success=False,
            errors=[ParseError(row=0, column="", message="未找到文件名包含'汇总'的 Excel 文件")]
        )

    logger.info(f"找到 Excel 文件: {excel_file}")

    # 2. 解析 Excel 并获取 image_count 映射
    parser = ExcelParser()
    result = parser.parse(str(excel_file))

    if not result.success:
        return result

    # 获取 image_count 映射
    image_count_map = _get_image_count_map(str(excel_file))
    logger.info(f"获取到 {len(image_count_map)} 个任务的图片数信息: {image_count_map}")

    # 3. 根据 image_count 精确匹配图片
    image_extensions = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}

    matched_count = 0
    for content_code, content in result.contents.items():
        # 清空原有的图片路径
        content.image_paths = []

        # 获取图片数
        image_count = image_count_map.get(content_code, 0)
        if not image_count or image_count <= 0:
            logger.warning(f"{content_code} 在Excel中的'图片数'列为空或0，跳过图片匹配")
            continue

        # 限制最多 9 张
        image_count = min(image_count, 9)

        # 按序号精确匹配图片: 文案编号 (1), 文案编号 (2), ...
        matched_images = []
        for i in range(1, image_count + 1):
            # 尝试多种文件名格式
            patterns = [
                f"{content_code} ({i})",      # F00621 (1)
                f"{content_code}({i})",       # F00621(1)
                f"{content_code}_{i}",        # F00621_1
                f"{content_code}-{i}",        # F00621-1
            ]

            found = False
            for img_file in folder.iterdir():
                if not img_file.is_file():
                    continue
                if img_file.suffix.lower() not in image_extensions:
                    continue

                # 检查文件名是否匹配任一模式
                stem = img_file.stem  # 不带扩展名的文件名
                for pattern in patterns:
                    if stem == pattern or stem.lower() == pattern.lower():
                        matched_images.append(str(img_file))
                        found = True
                        break
                if found:
                    break

            if not found:
                result.warnings.append(ParseWarning(
                    row=0, column="image_paths",
                    message=f"{content_code} 未找到第 {i} 张图片"
                ))

        content.image_paths = matched_images
        if matched_images:
            matched_count += 1
            logger.info(f"{content_code} 匹配到 {len(matched_images)}/{image_count} 张图片")

    logger.info(f"图片匹配完成: {matched_count}/{len(result.contents)} 个任务有图片")

    return result


def _get_image_count_map(excel_file: str) -> Dict[str, int]:
    """从 Excel 文件获取 content_code -> image_count 映射"""
    image_count_map = {}

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active

        # 解析表头
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        column_map = {}

        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            cell_str = str(cell_value).strip().lower()

            # 查找 content_code 和 image_count 列
            for col_name, aliases in [("content_code", COLUMN_MAPPINGS["content_code"]),
                                       ("image_count", COLUMN_MAPPINGS["image_count"])]:
                for alias in aliases:
                    if cell_str == alias.lower():
                        column_map[col_name] = idx
                        break

        if "content_code" not in column_map:
            logger.warning("Excel 中未找到文案编号列")
            wb.close()
            return image_count_map

        if "image_count" not in column_map:
            logger.warning("Excel 中未找到图片数列")
            wb.close()
            return image_count_map

        # 读取数据
        for row in ws.iter_rows(min_row=2, values_only=True):
            content_code_idx = column_map["content_code"]
            image_count_idx = column_map["image_count"]

            if content_code_idx < len(row) and image_count_idx < len(row):
                content_code = row[content_code_idx]
                image_count = row[image_count_idx]

                if content_code and image_count:
                    try:
                        image_count_map[str(content_code).strip()] = int(image_count)
                    except (ValueError, TypeError):
                        pass

        wb.close()

    except Exception as e:
        logger.error(f"读取 image_count 失败: {e}")

    return image_count_map


def _find_excel_file(folder: Path) -> Optional[Path]:
    """查找文件名包含'汇总'的 Excel 文件"""
    excel_extensions = {".xlsx", ".xls"}

    for f in folder.iterdir():
        if f.is_file() and f.suffix.lower() in excel_extensions:
            if "汇总" in f.name:
                return f

    return None
