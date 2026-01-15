"""Test cases for ExcelParser class (data/excel_parser.py)"""

import pytest
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from data.excel_parser import ExcelParser, ParseError, ParseWarning
from data.path_mapper import PathMapper
from models.enums import Channel, TaskStatus


@pytest.fixture
def parser(tmp_path):
    """Create an ExcelParser instance with temp directories"""
    share_root = tmp_path / "share"
    share_root.mkdir()
    cache_dir = tmp_path / "cache"
    cache_dir.mkdir()

    mapper = PathMapper(str(share_root), cache_dir)
    return ExcelParser(path_mapper=mapper)


@pytest.fixture
def valid_excel_file(tmp_path):
    """Create a valid Excel file for testing"""
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    file_path = tmp_path / "valid_schedule.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Write header
    headers = ["产品编码", "发布渠道", "排期时间", "产品名称", "文案内容", "图片路径", "群名"]
    ws.append(headers)

    # Write data rows
    data = [
        ["TEST001", "朋友圈", "2025-01-15 10:00", "Product A", "Test content 1", "image1.jpg", ""],
        ["TEST002", "群发", "2025-01-15 11:00", "Product B", "Test content 2", "image2.jpg", "Group1"],
        ["TEST003", "moment", "2025-01-15 12:00", "Product C", "Test content 3", "image3.jpg;image4.jpg", ""],
    ]

    for row in data:
        ws.append(row)

    wb.save(str(file_path))
    wb.close()

    return file_path


@pytest.fixture
def invalid_excel_file(tmp_path):
    """Create an Excel file with missing required columns"""
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl not available")

    file_path = tmp_path / "invalid_schedule.xlsx"

    wb = Workbook()
    ws = wb.active

    # Missing required columns
    headers = ["产品编码", "产品名称"]
    ws.append(headers)

    wb.save(str(file_path))
    wb.close()

    return file_path


# ==================== Parse Tests ====================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_schedule_file(parser, valid_excel_file):
    """Test parsing a valid schedule Excel file"""
    result = parser.parse(str(valid_excel_file))

    assert result.success is True
    assert len(result.tasks) == 3
    assert len(result.contents) == 3
    assert result.total_rows == 3
    assert result.valid_rows == 3

    # Check first task
    task1 = result.tasks[0]
    assert task1.content_code == "TEST001"
    assert task1.product_name == "Product A"
    assert task1.channel == Channel.moment
    assert task1.status == TaskStatus.pending

    # Check second task (group channel)
    task2 = result.tasks[1]
    assert task2.content_code == "TEST002"
    assert task2.channel == Channel.group
    assert task2.group_name == "Group1"

    # Check third task (multiple images)
    content3 = result.contents["TEST003"]
    assert len(content3.image_paths) == 2


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_invalid_file(parser, tmp_path):
    """Test parsing a non-existent file"""
    non_existent = tmp_path / "non_existent.xlsx"

    result = parser.parse(str(non_existent))

    assert result.success is False
    assert len(result.errors) > 0
    assert "不存在" in result.errors[0].message


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_missing_columns(parser, invalid_excel_file):
    """Test parsing file with missing required columns"""
    result = parser.parse(str(invalid_excel_file))

    assert result.success is False
    assert len(result.errors) > 0
    assert "缺少必需列" in result.errors[0].message


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_wrong_file_extension(parser, tmp_path):
    """Test parsing file with wrong extension"""
    wrong_file = tmp_path / "test.txt"
    wrong_file.write_text("not an excel file")

    result = parser.parse(str(wrong_file))

    assert result.success is False
    assert len(result.errors) > 0
    assert "不支持的文件格式" in result.errors[0].message


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_with_empty_rows(parser, tmp_path):
    """Test parsing file with empty rows"""
    file_path = tmp_path / "with_empty_rows.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间"]
    ws.append(headers)
    ws.append(["TEST001", "朋友圈", "2025-01-15 10:00"])
    ws.append([None, None, None])  # Empty row
    ws.append(["TEST002", "群发", "2025-01-15 11:00"])

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert result.success is True
    assert len(result.tasks) == 2  # Empty row should be skipped


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_invalid_channel(parser, tmp_path):
    """Test parsing with invalid channel value"""
    file_path = tmp_path / "invalid_channel.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间"]
    ws.append(headers)
    ws.append(["TEST001", "invalid_channel", "2025-01-15 10:00"])

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert len(result.errors) > 0
    assert any("无效的渠道值" in err.message for err in result.errors)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_invalid_datetime(parser, tmp_path):
    """Test parsing with invalid datetime format"""
    file_path = tmp_path / "invalid_datetime.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间"]
    ws.append(headers)
    ws.append(["TEST001", "朋友圈", "not a date"])

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert len(result.errors) > 0
    assert any("无效的时间格式" in err.message for err in result.errors)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_strict_mode(parser, tmp_path):
    """Test parsing in strict mode stops on first error"""
    parser.strict_mode = True

    file_path = tmp_path / "with_errors.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间"]
    ws.append(headers)
    ws.append(["TEST001", "invalid", "2025-01-15 10:00"])  # Error
    ws.append(["TEST002", "朋友圈", "2025-01-15 11:00"])  # Valid

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert result.success is False
    assert len(result.tasks) == 0  # No tasks should be created


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_multiple_image_paths(parser, tmp_path):
    """Test parsing multiple image paths with different separators"""
    file_path = tmp_path / "multi_images.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间", "图片路径"]
    ws.append(headers)
    ws.append(["TEST001", "朋友圈", "2025-01-15 10:00", "img1.jpg;img2.jpg;img3.jpg"])
    ws.append(["TEST002", "朋友圈", "2025-01-15 11:00", "img4.jpg\nimg5.jpg"])

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert result.success is True

    content1 = result.contents["TEST001"]
    assert len(content1.image_paths) == 3

    content2 = result.contents["TEST002"]
    assert len(content2.image_paths) == 2


# ==================== Channel Parsing Tests ====================

def test_parse_channel_moment_variants(parser):
    """Test parsing different moment channel values"""
    variants = ["moment", "朋友圈", "pyq", "moments"]

    for variant in variants:
        channel = parser._parse_channel(variant)
        assert channel == Channel.moment


def test_parse_channel_group_variants(parser):
    """Test parsing different group channel values"""
    variants = ["group", "群发", "qf", "群", "groups"]

    for variant in variants:
        channel = parser._parse_channel(variant)
        assert channel == Channel.group


def test_parse_channel_invalid(parser):
    """Test parsing invalid channel returns None"""
    channel = parser._parse_channel("invalid")
    assert channel is None


# ==================== DateTime Parsing Tests ====================

def test_parse_datetime_formats(parser):
    """Test parsing various datetime formats"""
    formats = [
        ("2025-01-15 10:30:00", datetime(2025, 1, 15, 10, 30, 0)),
        ("2025-01-15 10:30", datetime(2025, 1, 15, 10, 30)),
        ("2025/01/15 10:30:00", datetime(2025, 1, 15, 10, 30, 0)),
        ("2025.01.15 10:30", datetime(2025, 1, 15, 10, 30)),
        ("2025-01-15", datetime(2025, 1, 15, 0, 0)),
    ]

    for date_str, expected in formats:
        result = parser._parse_datetime(date_str)
        assert result == expected


def test_parse_datetime_already_datetime(parser):
    """Test parsing datetime that's already a datetime object"""
    dt = datetime(2025, 1, 15, 10, 30)
    result = parser._parse_datetime(dt)

    assert result == dt


def test_parse_datetime_invalid(parser):
    """Test parsing invalid datetime returns None"""
    result = parser._parse_datetime("not a date")
    assert result is None


# ==================== Image Path Parsing Tests ====================

def test_parse_image_paths_semicolon(parser):
    """Test parsing semicolon-separated image paths"""
    paths = parser._parse_image_paths("img1.jpg;img2.jpg;img3.jpg")

    assert len(paths) == 3
    assert "img1.jpg" in paths[0]
    assert "img2.jpg" in paths[1]


def test_parse_image_paths_newline(parser):
    """Test parsing newline-separated image paths"""
    paths = parser._parse_image_paths("img1.jpg\nimg2.jpg\nimg3.jpg")

    assert len(paths) == 3


def test_parse_image_paths_empty(parser):
    """Test parsing empty image paths"""
    paths = parser._parse_image_paths("")
    assert len(paths) == 0

    paths = parser._parse_image_paths(None)
    assert len(paths) == 0


# ==================== Validation Tests ====================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_validate_file_valid(parser, valid_excel_file):
    """Test validating a valid Excel file"""
    is_valid, errors = parser.validate_file(str(valid_excel_file))

    assert is_valid is True
    assert len(errors) == 0


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_validate_file_missing_columns(parser, invalid_excel_file):
    """Test validating file with missing columns"""
    is_valid, errors = parser.validate_file(str(invalid_excel_file))

    assert is_valid is False
    assert len(errors) > 0


def test_validate_file_not_exists(parser, tmp_path):
    """Test validating non-existent file"""
    non_existent = tmp_path / "non_existent.xlsx"

    is_valid, errors = parser.validate_file(str(non_existent))

    assert is_valid is False
    assert len(errors) > 0


# ==================== Content Merging Tests ====================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_merge_duplicate_content_codes(parser, tmp_path):
    """Test that duplicate content codes merge images"""
    file_path = tmp_path / "duplicate_codes.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间", "文案内容", "图片路径"]
    ws.append(headers)
    ws.append(["TEST001", "朋友圈", "2025-01-15 10:00", "Content 1", "img1.jpg"])
    ws.append(["TEST001", "朋友圈", "2025-01-15 11:00", "", "img2.jpg"])  # Same code

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert result.success is True
    assert len(result.tasks) == 2  # Two tasks created
    assert len(result.contents) == 1  # But only one content (merged)

    # Check merged content has both images
    content = result.contents["TEST001"]
    assert len(content.image_paths) == 2


# ==================== Warning Tests ====================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_group_without_group_name(parser, tmp_path):
    """Test warning when group channel has no group name"""
    file_path = tmp_path / "group_no_name.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间", "群名"]
    ws.append(headers)
    ws.append(["TEST001", "群发", "2025-01-15 10:00", ""])  # No group name

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert result.success is True
    assert len(result.warnings) > 0
    assert any("建议填写群名" in warn.message for warn in result.warnings)


# ==================== Edge Cases ====================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_empty_content_code(parser, tmp_path):
    """Test error when content code is empty"""
    file_path = tmp_path / "empty_code.xlsx"

    wb = Workbook()
    ws = wb.active

    headers = ["产品编码", "发布渠道", "排期时间"]
    ws.append(headers)
    ws.append(["", "朋友圈", "2025-01-15 10:00"])  # Empty content code

    wb.save(str(file_path))
    wb.close()

    result = parser.parse(str(file_path))

    assert len(result.errors) > 0
    assert any("不能为空" in err.message for err in result.errors)


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
def test_parse_with_sheet_name(parser, tmp_path):
    """Test parsing specific sheet by name"""
    file_path = tmp_path / "multi_sheet.xlsx"

    wb = Workbook()

    # Create first sheet
    ws1 = wb.active
    ws1.title = "Sheet1"
    headers = ["产品编码", "发布渠道", "排期时间"]
    ws1.append(headers)
    ws1.append(["TEST001", "朋友圈", "2025-01-15 10:00"])

    # Create second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(headers)
    ws2.append(["TEST002", "群发", "2025-01-15 11:00"])

    wb.save(str(file_path))
    wb.close()

    # Parse specific sheet
    result = parser.parse(str(file_path), sheet_name="Sheet2")

    assert result.success is True
    assert len(result.tasks) == 1
    assert result.tasks[0].content_code == "TEST002"
